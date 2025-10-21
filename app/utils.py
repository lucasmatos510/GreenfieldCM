from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from app.models import RegistroHora, Funcionario
import tempfile
import os

def gerar_relatorio_excel(tipo='mensal', funcionario_id=None, mes=None, ano=None):
    """
    Gera relatório Excel de horas trabalhadas
    
    Args:
        tipo (str): 'diario', 'mensal' ou 'anual'
        funcionario_id (int): ID do funcionário (opcional)
        mes (int): Mês para relatório mensal
        ano (int): Ano para relatório
    
    Returns:
        str: Caminho do arquivo Excel gerado
    """
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    
    # Configurar estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="366092")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if tipo == 'mensal':
        ws.title = "Relatório Mensal"
        dados = _gerar_dados_mensais(funcionario_id, mes, ano)
        _criar_planilha_mensal(ws, dados, header_font, header_fill, center_alignment, border)
    
    elif tipo == 'anual':
        ws.title = "Relatório Anual"
        dados = _gerar_dados_anuais(funcionario_id, ano)
        _criar_planilha_anual(ws, dados, header_font, header_fill, center_alignment, border)
    
    elif tipo == 'diario':
        ws.title = "Relatório Diário"
        dados = _gerar_dados_diarios(funcionario_id, mes, ano)
        _criar_planilha_diaria(ws, dados, header_font, header_fill, center_alignment, border)
    
    # Criar arquivo temporário para envio
    import tempfile
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    # Também salvar na pasta Downloads
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
    if not os.path.exists(downloads_path):
        os.makedirs(downloads_path)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"relatorio_horas_{tipo}_{timestamp}.xlsx"
    caminho_downloads = os.path.join(downloads_path, nome_arquivo)
    
    # Copiar para Downloads
    import shutil
    shutil.copy2(temp_file.name, caminho_downloads)
    
    return temp_file.name

def _gerar_dados_mensais(funcionario_id=None, mes=None, ano=None):
    """Gera dados para relatório mensal"""
    if not mes:
        mes = datetime.now().month
    if not ano:
        ano = datetime.now().year
    
    query = RegistroHora.query.filter(
        RegistroHora.data >= date(ano, mes, 1)
    )
    
    # Último dia do mês
    if mes == 12:
        ultimo_dia = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        ultimo_dia = date(ano, mes + 1, 1) - timedelta(days=1)
    
    query = query.filter(RegistroHora.data <= ultimo_dia)
    
    if funcionario_id:
        query = query.filter(RegistroHora.funcionario_id == funcionario_id)
    
    registros = query.order_by(RegistroHora.funcionario_id, RegistroHora.data).all()
    
    # Agrupar por funcionário
    dados_funcionarios = {}
    for registro in registros:
        func_id = registro.funcionario_id
        if func_id not in dados_funcionarios:
            dados_funcionarios[func_id] = {
                'funcionario': registro.funcionario,
                'registros': [],
                'total_horas': 0
            }
        
        dados_funcionarios[func_id]['registros'].append(registro)
        dados_funcionarios[func_id]['total_horas'] += registro.horas
    
    return {
        'funcionarios': dados_funcionarios,
        'periodo': f"{mes:02d}/{ano}",
        'tipo': 'mensal'
    }

def _gerar_dados_anuais(funcionario_id=None, ano=None):
    """Gera dados para relatório anual"""
    if not ano:
        ano = datetime.now().year
    
    query = RegistroHora.query.filter(
        RegistroHora.data >= date(ano, 1, 1),
        RegistroHora.data <= date(ano, 12, 31)
    )
    
    if funcionario_id:
        query = query.filter(RegistroHora.funcionario_id == funcionario_id)
    
    registros = query.order_by(RegistroHora.funcionario_id, RegistroHora.data).all()
    
    # Agrupar por funcionário e mês
    dados_funcionarios = {}
    for registro in registros:
        func_id = registro.funcionario_id
        mes = registro.data.month
        
        if func_id not in dados_funcionarios:
            dados_funcionarios[func_id] = {
                'funcionario': registro.funcionario,
                'meses': {i: {'horas': 0, 'registros': 0} for i in range(1, 13)},
                'total_horas': 0
            }
        
        dados_funcionarios[func_id]['meses'][mes]['horas'] += registro.horas
        dados_funcionarios[func_id]['meses'][mes]['registros'] += 1
        dados_funcionarios[func_id]['total_horas'] += registro.horas
    
    return {
        'funcionarios': dados_funcionarios,
        'ano': ano,
        'tipo': 'anual'
    }

def _gerar_dados_diarios(funcionario_id=None, mes=None, ano=None):
    """Gera dados para relatório diário"""
    if not mes:
        mes = datetime.now().month
    if not ano:
        ano = datetime.now().year
    
    query = RegistroHora.query.filter(
        RegistroHora.data >= date(ano, mes, 1)
    )
    
    # Último dia do mês
    if mes == 12:
        ultimo_dia = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        ultimo_dia = date(ano, mes + 1, 1) - timedelta(days=1)
    
    query = query.filter(RegistroHora.data <= ultimo_dia)
    
    if funcionario_id:
        query = query.filter(RegistroHora.funcionario_id == funcionario_id)
    
    registros = query.order_by(RegistroHora.data, RegistroHora.funcionario_id).all()
    
    return {
        'registros': registros,
        'periodo': f"{mes:02d}/{ano}",
        'tipo': 'diario'
    }

def _criar_planilha_mensal(ws, dados, header_font, header_fill, center_alignment, border):
    """Cria planilha de relatório mensal no formato solicitado"""
    # Cabeçalho principal
    ws.merge_cells('A1:D1')
    ws['A1'] = f"RELATÓRIO MENSAL DE HORAS - {dados['periodo']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment
    
    # Cabeçalhos das colunas
    headers = ['Área de Atuação', 'Cargo', 'Nome do Funcionário', 'Horas']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Dados dos funcionários - ordenados por área e cargo
    row = 4
    total_geral_horas = 0
    
    # Criar lista ordenada de funcionários
    funcionarios_ordenados = []
    for func_data in dados['funcionarios'].values():
        funcionario = func_data['funcionario']
        area_nome = funcionario.cargo.area.nome if funcionario.cargo and funcionario.cargo.area else 'N/A'
        cargo_nome = funcionario.cargo.nome if funcionario.cargo else 'N/A'
        
        funcionarios_ordenados.append({
            'area': area_nome,
            'cargo': cargo_nome,
            'funcionario': funcionario,
            'horas': func_data['total_horas']
        })
    
    # Ordenar por área e depois por cargo
    funcionarios_ordenados.sort(key=lambda x: (x['area'], x['cargo'], x['funcionario'].nome))
    
    # Preencher dados na planilha
    for func_info in funcionarios_ordenados:
        total_horas = round(func_info['horas'], 2)
        
        # Só adicionar linha se funcionário tem horas registradas
        if total_horas > 0:
            ws.cell(row=row, column=1, value=func_info['area']).border = border
            ws.cell(row=row, column=2, value=func_info['cargo']).border = border
            ws.cell(row=row, column=3, value=func_info['funcionario'].nome).border = border
            ws.cell(row=row, column=4, value=total_horas).border = border
            
            total_geral_horas += total_horas
            row += 1
    
    # Total geral
    if row > 4:  # Se há dados
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "TOTAL GERAL"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        ws[f'D{row}'] = round(total_geral_horas, 2)
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'D{row}'].border = border
    
    # Ajustar largura das colunas
    larguras = [25, 20, 25, 15]  # Área, Cargo, Nome, Horas
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = largura

def _criar_planilha_anual(ws, dados, header_font, header_fill, center_alignment, border):
    """Cria planilha de relatório anual organizado por área e cargo"""
    # Cabeçalho principal
    ws.merge_cells('A1:P1')
    ws['A1'] = f"RELATÓRIO ANUAL DE HORAS - {dados['ano']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment
    
    # Cabeçalhos das colunas
    meses = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
             'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    headers = ['Área de Atuação', 'Cargo', 'Nome do Funcionário'] + meses + ['Total']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Organizar funcionários por área e cargo
    funcionarios_ordenados = []
    for func_data in dados['funcionarios'].values():
        funcionario = func_data['funcionario']
        area_nome = funcionario.cargo.area.nome if funcionario.cargo and funcionario.cargo.area else 'N/A'
        cargo_nome = funcionario.cargo.nome if funcionario.cargo else 'N/A'
        
        funcionarios_ordenados.append({
            'area': area_nome,
            'cargo': cargo_nome,
            'funcionario': funcionario,
            'meses': func_data['meses'],
            'total_horas': func_data['total_horas']
        })
    
    # Ordenar por área e depois por cargo
    funcionarios_ordenados.sort(key=lambda x: (x['area'], x['cargo'], x['funcionario'].nome))
    
    # Dados dos funcionários
    row = 4
    for func_info in funcionarios_ordenados:
        # Só incluir funcionários com horas registradas
        if func_info['total_horas'] > 0:
            ws.cell(row=row, column=1, value=func_info['area']).border = border
            ws.cell(row=row, column=2, value=func_info['cargo']).border = border
            ws.cell(row=row, column=3, value=func_info['funcionario'].nome).border = border
            
            for mes in range(1, 13):
                horas = round(func_info['meses'][mes]['horas'], 2)
                ws.cell(row=row, column=mes + 3, value=horas if horas > 0 else '-').border = border
            
            total = round(func_info['total_horas'], 2)
            ws.cell(row=row, column=16, value=total).border = border
            row += 1
    
    # Ajustar largura das colunas
    larguras = [25, 20, 25] + [10] * 12 + [12]  # Área, Cargo, Nome, 12 meses, Total
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = largura

def _criar_planilha_diaria(ws, dados, header_font, header_fill, center_alignment, border):
    """Cria planilha de relatório diário organizado por área e cargo"""
    # Cabeçalho principal  
    ws.merge_cells('A1:E1')
    ws['A1'] = f"RELATÓRIO DIÁRIO DE HORAS - {dados['periodo']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment
    
    # Cabeçalhos das colunas
    headers = ['Área de Atuação', 'Cargo', 'Nome do Funcionário', 'Data', 'Horas']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Organizar registros por área/cargo/funcionário
    registros_organizados = []
    for registro in dados['registros']:
        area_nome = registro.funcionario.cargo.area.nome if registro.funcionario.cargo and registro.funcionario.cargo.area else 'N/A'
        cargo_nome = registro.funcionario.cargo.nome if registro.funcionario.cargo else 'N/A'
        
        registros_organizados.append({
            'area': area_nome,
            'cargo': cargo_nome,
            'funcionario': registro.funcionario.nome,
            'data': registro.data.strftime('%d/%m/%Y'),
            'horas': round(registro.horas, 2)
        })
    
    # Ordenar por área, cargo, funcionário e data
    registros_organizados.sort(key=lambda x: (x['area'], x['cargo'], x['funcionario'], x['data']))
    
    # Preencher dados na planilha
    row = 4
    total_horas = 0
    
    for reg in registros_organizados:
        ws.cell(row=row, column=1, value=reg['area']).border = border
        ws.cell(row=row, column=2, value=reg['cargo']).border = border
        ws.cell(row=row, column=3, value=reg['funcionario']).border = border
        ws.cell(row=row, column=4, value=reg['data']).border = border
        ws.cell(row=row, column=5, value=reg['horas']).border = border
        
        total_horas += reg['horas']
        row += 1
    
    # Total geral
    if row > 4:  # Se há dados
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "TOTAL GERAL"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        ws[f'E{row}'] = round(total_horas, 2)
        ws[f'E{row}'].font = Font(bold=True)
        ws[f'E{row}'].border = border
    
    # Ajustar largura das colunas
    larguras = [25, 20, 25, 15, 15]  # Área, Cargo, Nome, Data, Horas
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = largura